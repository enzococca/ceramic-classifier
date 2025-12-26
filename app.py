#!/usr/bin/env python3
"""
Ceramica Classifier - Universal Pottery Classification System
Supports multiple database types with AI-powered schema analysis
"""

from flask import Flask, render_template, jsonify, request, send_file, session
from flask_socketio import SocketIO, emit
import os
import json
import base64
import requests
import threading
from pathlib import Path
from PIL import Image
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill
import tempfile
import traceback

from db_connector import create_connector, test_connection, CONNECTOR_TYPES
from ai_analyzer import AISchemaAnalyzer, test_api_key
from plate_generator import PlateGenerator, get_layouts, get_layout_by_id, generate_excel_report

app = Flask(__name__)
app.config['SECRET_KEY'] = 'ceramica-classifier-universal-2024'
socketio = SocketIO(app, cors_allowed_origins="*", async_mode='eventlet')

# Directories
BASE_DIR = Path(__file__).parent
CONFIG_DIR = BASE_DIR / "configs"
EXPORT_DIR = BASE_DIR / "exports"
CONFIG_DIR.mkdir(exist_ok=True)
EXPORT_DIR.mkdir(exist_ok=True)

# Default ML API URL
ML_API_URL = os.environ.get('ML_API_URL', 'https://pottery-comparison-oman.up.railway.app/api/ml/similar')

# Global state
app_state = {
    'db_connector': None,
    'config': None,
    'api_key': None,
    'classification': {
        'running': False,
        'paused': False,
        'total': 0,
        'processed': 0,
        'errors': 0,
        'results': [],
        'statistics': {
            'periods': {},
            'decorations': {},
            'confidences': [],
            'sites': {}
        }
    }
}


# ============== Configuration Routes ==============

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/api/db-types')
def get_db_types():
    """Get supported database types."""
    return jsonify({
        'types': [
            {'id': 'postgresql', 'name': 'PostgreSQL', 'icon': '🐘'},
            {'id': 'mysql', 'name': 'MySQL', 'icon': '🐬'},
            {'id': 'sqlite', 'name': 'SQLite', 'icon': '📦'},
            {'id': 'mongodb', 'name': 'MongoDB', 'icon': '🍃'},
            {'id': 'excel', 'name': 'Excel', 'icon': '📊'},
            {'id': 'csv', 'name': 'CSV', 'icon': '📄'},
        ]
    })


@app.route('/api/test-connection', methods=['POST'])
def api_test_connection():
    """Test database connection."""
    data = request.json
    db_type = data.get('db_type')
    config = data.get('config', {})

    result = test_connection(db_type, config)
    return jsonify(result)


@app.route('/api/get-schema', methods=['POST'])
def api_get_schema():
    """Get database schema information."""
    data = request.json
    db_type = data.get('db_type')
    config = data.get('config', {})

    try:
        connector = create_connector(db_type, config)
        if connector.connect():
            schema = connector.get_schema_info()
            connector.disconnect()
            return jsonify({'success': True, 'schema': schema})
        return jsonify({'success': False, 'error': 'Connection failed'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/test-api-key', methods=['POST'])
def api_test_api_key():
    """Test Claude API key."""
    data = request.json
    api_key = data.get('api_key')

    result = test_api_key(api_key)
    if result.get('valid'):
        app_state['api_key'] = api_key
    return jsonify(result)


@app.route('/api/analyze-schema', methods=['POST'])
def api_analyze_schema():
    """Use AI to analyze database schema and suggest configuration."""
    data = request.json
    api_key = data.get('api_key') or app_state.get('api_key')
    schema = data.get('schema')
    image_path = data.get('image_path', '')

    if not api_key:
        return jsonify({'success': False, 'error': 'API key not provided'})

    if not schema:
        return jsonify({'success': False, 'error': 'Schema not provided'})

    analyzer = AISchemaAnalyzer(api_key)
    result = analyzer.analyze_schema(schema, image_path)

    if result.get('success'):
        # Generate query
        query = analyzer.generate_query(result)
        result['generated_query'] = query

    return jsonify(result)


@app.route('/api/save-config', methods=['POST'])
def api_save_config():
    """Save configuration to file."""
    data = request.json
    config_name = data.get('name', 'default')
    config = data.get('config')

    try:
        config_path = CONFIG_DIR / f"{config_name}.json"
        with open(config_path, 'w') as f:
            json.dump(config, f, indent=2)

        app_state['config'] = config
        return jsonify({'success': True, 'path': str(config_path)})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/load-config/<name>')
def api_load_config(name):
    """Load configuration from file."""
    try:
        config_path = CONFIG_DIR / f"{name}.json"
        with open(config_path) as f:
            config = json.load(f)
        app_state['config'] = config
        return jsonify({'success': True, 'config': config})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/list-configs')
def api_list_configs():
    """List saved configurations."""
    configs = []
    for f in CONFIG_DIR.glob('*.json'):
        configs.append({
            'name': f.stem,
            'path': str(f),
            'modified': datetime.fromtimestamp(f.stat().st_mtime).isoformat()
        })
    return jsonify({'configs': configs})


# ============== API Key Management ==============

API_KEY_FILE = BASE_DIR / ".api_key"

@app.route('/api/save-api-key', methods=['POST'])
def api_save_api_key():
    """Save API key to file."""
    data = request.json
    api_key = data.get('api_key', '')

    try:
        with open(API_KEY_FILE, 'w') as f:
            f.write(api_key)
        return jsonify({'success': True, 'message': 'API key saved'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/load-api-key')
def api_load_api_key():
    """Load saved API key."""
    try:
        if API_KEY_FILE.exists():
            with open(API_KEY_FILE, 'r') as f:
                api_key = f.read().strip()
            # Mask the key for display (show only last 4 chars)
            masked = '•' * (len(api_key) - 4) + api_key[-4:] if len(api_key) > 4 else '•' * len(api_key)
            return jsonify({'success': True, 'api_key': api_key, 'masked': masked})
        return jsonify({'success': False, 'message': 'No saved API key'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


# ============== File Browser ==============

@app.route('/api/browse')
def api_browse():
    """Browse filesystem directories."""
    path = request.args.get('path', os.path.expanduser('~'))

    try:
        path = Path(path)
        if not path.exists():
            path = Path.home()

        items = []

        # Add parent directory option
        if path.parent != path:
            items.append({
                'name': '..',
                'path': str(path.parent),
                'is_dir': True,
                'type': 'directory',
                'icon': '📁'
            })

        # List directory contents
        for item in sorted(path.iterdir(), key=lambda x: (not x.is_dir(), x.name.lower())):
            if item.name.startswith('.'):
                continue  # Skip hidden files

            if item.is_dir():
                items.append({
                    'name': item.name,
                    'path': str(item),
                    'is_dir': True,
                    'type': 'directory',
                    'icon': '📁'
                })
            else:
                # Check if it's a supported file type
                ext = item.suffix.lower()
                if ext in ['.xlsx', '.xls', '.csv', '.db', '.sqlite', '.sqlite3']:
                    icon = '📊' if ext in ['.xlsx', '.xls'] else ('📄' if ext == '.csv' else '🗃️')
                    try:
                        size = item.stat().st_size
                    except:
                        size = 0
                    items.append({
                        'name': item.name,
                        'path': str(item),
                        'is_dir': False,
                        'type': 'file',
                        'icon': icon,
                        'size': size
                    })

        return jsonify({
            'success': True,
            'current_path': str(path),
            'items': items
        })
    except PermissionError:
        return jsonify({'success': False, 'error': 'Permission denied'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/browse-images')
def api_browse_images():
    """Browse filesystem for image directories."""
    path = request.args.get('path', os.path.expanduser('~'))

    try:
        path = Path(path)
        if not path.exists():
            path = Path.home()

        items = []
        image_count = 0

        # Add parent directory
        if path.parent != path:
            items.append({
                'name': '..',
                'path': str(path.parent),
                'is_dir': True,
                'type': 'directory',
                'icon': '📁'
            })

        # List directories and count images
        for item in sorted(path.iterdir(), key=lambda x: (not x.is_dir(), x.name.lower())):
            if item.name.startswith('.'):
                continue

            if item.is_dir():
                # Count images in subdirectory
                try:
                    img_count = sum(1 for f in item.iterdir()
                                   if f.suffix.lower() in ['.jpg', '.jpeg', '.png', '.tif', '.tiff'])
                except:
                    img_count = 0

                items.append({
                    'name': item.name,
                    'path': str(item),
                    'is_dir': True,
                    'type': 'directory',
                    'icon': '📁',
                    'image_count': img_count
                })
            elif item.suffix.lower() in ['.jpg', '.jpeg', '.png', '.tif', '.tiff']:
                image_count += 1

        return jsonify({
            'success': True,
            'current_path': str(path),
            'items': items,
            'image_count': image_count
        })
    except PermissionError:
        return jsonify({'success': False, 'error': 'Permission denied'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/docs/<lang>')
def api_docs(lang):
    """Serve documentation in specified language."""
    try:
        # Define doc files for each language
        docs_dir = BASE_DIR / 'docs'

        if lang == 'en':
            files = ['QUICK_START_EN.md', 'DATABASE_STRUCTURE_EN.md']
        else:  # Italian
            files = ['QUICK_START.md', 'DATABASE_STRUCTURE.md']

        content_parts = []
        for filename in files:
            filepath = docs_dir / filename
            if filepath.exists():
                with open(filepath, 'r', encoding='utf-8') as f:
                    content_parts.append(f.read())

        if not content_parts:
            return jsonify({'success': False, 'error': 'Documentation not found'})

        markdown_content = '\n\n---\n\n'.join(content_parts)

        # Convert markdown to HTML (basic conversion)
        html = convert_markdown_to_html(markdown_content)

        return jsonify({'success': True, 'html': html})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


def convert_markdown_to_html(md_text):
    """Convert markdown to HTML with basic formatting."""
    import re

    html = md_text

    # Escape HTML special chars first (except in code blocks)
    # We'll handle this more carefully

    # Code blocks (```...```)
    def replace_code_block(match):
        lang = match.group(1) or ''
        code = match.group(2)
        # Escape HTML in code
        code = code.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
        return f'<pre><code class="language-{lang}">{code}</code></pre>'

    html = re.sub(r'```(\w*)\n(.*?)```', replace_code_block, html, flags=re.DOTALL)

    # Inline code
    html = re.sub(r'`([^`]+)`', r'<code>\1</code>', html)

    # Headers
    html = re.sub(r'^### (.+)$', r'<h3>\1</h3>', html, flags=re.MULTILINE)
    html = re.sub(r'^## (.+)$', r'<h2>\1</h2>', html, flags=re.MULTILINE)
    html = re.sub(r'^# (.+)$', r'<h1>\1</h1>', html, flags=re.MULTILINE)

    # Bold and italic
    html = re.sub(r'\*\*([^*]+)\*\*', r'<strong>\1</strong>', html)
    html = re.sub(r'\*([^*]+)\*', r'<em>\1</em>', html)

    # Links
    html = re.sub(r'\[([^\]]+)\]\(([^)]+)\)', r'<a href="\2" target="_blank">\1</a>', html)

    # Tables
    def replace_table(match):
        lines = match.group(0).strip().split('\n')
        if len(lines) < 2:
            return match.group(0)

        result = '<table>'
        for i, line in enumerate(lines):
            if '---' in line:
                continue
            cells = [c.strip() for c in line.strip('|').split('|')]
            tag = 'th' if i == 0 else 'td'
            result += '<tr>' + ''.join(f'<{tag}>{c}</{tag}>' for c in cells) + '</tr>'
        result += '</table>'
        return result

    html = re.sub(r'(\|.+\|[\n])+', replace_table, html)

    # Lists
    html = re.sub(r'^- (.+)$', r'<li>\1</li>', html, flags=re.MULTILINE)
    html = re.sub(r'(<li>.*</li>\n?)+', r'<ul>\g<0></ul>', html)

    # Numbered lists
    html = re.sub(r'^\d+\. (.+)$', r'<li>\1</li>', html, flags=re.MULTILINE)

    # Horizontal rules
    html = re.sub(r'^---+$', r'<hr>', html, flags=re.MULTILINE)

    # Paragraphs (simple)
    html = re.sub(r'\n\n+', '</p><p>', html)
    html = '<p>' + html + '</p>'

    # Clean up empty paragraphs
    html = re.sub(r'<p>\s*</p>', '', html)
    html = re.sub(r'<p>\s*(<h[123])', r'\1', html)
    html = re.sub(r'(</h[123]>)\s*</p>', r'\1', html)
    html = re.sub(r'<p>\s*(<pre)', r'\1', html)
    html = re.sub(r'(</pre>)\s*</p>', r'\1', html)
    html = re.sub(r'<p>\s*(<ul)', r'\1', html)
    html = re.sub(r'(</ul>)\s*</p>', r'\1', html)
    html = re.sub(r'<p>\s*(<table)', r'\1', html)
    html = re.sub(r'(</table>)\s*</p>', r'\1', html)
    html = re.sub(r'<p>\s*(<hr)', r'\1', html)
    html = re.sub(r'(<hr>)\s*</p>', r'\1', html)

    return html


# ============== Classification Routes ==============

def find_image_file(base_path: str, pattern: str, item: dict) -> Path:
    """Find image file using pattern and item data."""
    base = Path(base_path)

    # If pattern contains full path, extract just the filename pattern
    if '/' in pattern and '{' in pattern:
        pattern = pattern.split('/')[-1]

    # Replace placeholders in pattern
    filename = pattern
    for key, value in item.items():
        if value is not None:
            filename = filename.replace(f'{{{key}}}', str(value))

    # Remove any remaining unresolved placeholders
    import re
    filename = re.sub(r'\{[^}]+\}', '', filename)

    # Try direct path with pattern
    file_path = base / filename
    if file_path.exists():
        return file_path

    # Try with common extensions
    base_name = filename.rsplit('.', 1)[0] if '.' in filename else filename
    for ext in ['.png', '.jpg', '.jpeg', '.PNG', '.JPG', '.JPEG', '.tif', '.TIF']:
        test_path = base / f"{base_name}{ext}"
        if test_path.exists():
            return test_path

    # Try common patterns for media files: {id_media}_{filename}.ext
    media_id = item.get('media_id')
    fname = item.get('filename', '')

    if media_id:
        # Pattern: {media_id}_{filename}.png (common in archaeological databases)
        for ext in ['.png', '.jpg', '.jpeg', '.PNG', '.JPG', '.JPEG']:
            if fname:
                test_path = base / f"{media_id}_{fname}{ext}"
                if test_path.exists():
                    return test_path
            # Just media_id
            test_path = base / f"{media_id}{ext}"
            if test_path.exists():
                return test_path

    # Try glob pattern as fallback
    search_term = str(media_id) if media_id else fname
    if search_term:
        for f in base.glob(f"*{search_term}*"):
            if not f.name.startswith('._') and f.is_file():
                return f

    return None


def create_thumbnail_base64(image_path, size=(150, 150)):
    """Create base64 thumbnail."""
    try:
        with Image.open(image_path) as img:
            img.thumbnail(size, Image.Resampling.LANCZOS)
            if img.mode in ('RGBA', 'P'):
                img = img.convert('RGB')
            buffer = io.BytesIO()
            img.save(buffer, format='JPEG', quality=80)
            b64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
        return f"data:image/jpeg;base64,{b64}"
    except:
        return None


def image_to_base64(image_path, max_size=800):
    """Convert image to base64 for ML API."""
    with Image.open(image_path) as img:
        if max(img.size) > max_size:
            img.thumbnail((max_size, max_size), Image.Resampling.LANCZOS)
        if img.mode in ('RGBA', 'P'):
            img = img.convert('RGB')
        buffer = io.BytesIO()
        img.save(buffer, format='JPEG', quality=85)
        b64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
    return f"data:image/jpeg;base64,{b64}"


def classify_image(image_path, top_k=5):
    """Classify image using ML API."""
    try:
        image_data = image_to_base64(image_path)
        print(f"[ML] Calling API for: {image_path}")

        response = requests.post(
            ML_API_URL,
            json={'image': image_data, 'top_k': top_k},
            timeout=120
        )

        print(f"[ML] Response status: {response.status_code}, length: {len(response.text)}")

        if response.status_code == 200:
            # Check if response is empty
            if not response.text or response.text.strip() == '':
                return {'success': False, 'error': 'Empty response from ML API'}

            try:
                result = response.json()
            except Exception as json_err:
                print(f"[ML] JSON parse error: {json_err}, response text: {response.text[:200]}")
                return {'success': False, 'error': f'JSON parse error: {str(json_err)}'}

            if result.get('success') and result.get('analysis'):
                return {
                    'success': True,
                    'period_suggestion': result['analysis'].get('period_suggestion', ''),
                    'period_confidence': result['analysis'].get('period_confidence', 0),
                    'decoration_suggestion': result['analysis'].get('decoration_suggestion', ''),
                    'sites': result['analysis'].get('sites', []),
                    'top_match': result['similar_items'][0] if result.get('similar_items') else None,
                    'similar_items': result.get('similar_items', [])[:5],
                    'analysis_text': result['analysis'].get('text', '')
                }
            else:
                error_msg = result.get('error', 'Unknown API error')
                return {'success': False, 'error': f"API returned: {error_msg}"}
        else:
            return {'success': False, 'error': f"API error: {response.status_code} - {response.text[:100]}"}
    except requests.exceptions.Timeout:
        return {'success': False, 'error': 'ML API timeout (120s)'}
    except requests.exceptions.ConnectionError as e:
        return {'success': False, 'error': f'Connection error: {str(e)[:100]}'}
    except Exception as e:
        print(f"[ML] Exception: {type(e).__name__}: {str(e)}")
        return {'success': False, 'error': str(e)}


def update_statistics(classification):
    """Update classification statistics."""
    if classification.get('success'):
        stats = app_state['classification']['statistics']

        period = classification.get('period_suggestion', 'Unknown')
        decoration = classification.get('decoration_suggestion', 'Unknown')
        confidence = classification.get('period_confidence', 0)

        stats['periods'][period] = stats['periods'].get(period, 0) + 1
        stats['decorations'][decoration] = stats['decorations'].get(decoration, 0) + 1
        stats['confidences'].append(confidence)

        for site in classification.get('sites', []):
            stats['sites'][site] = stats['sites'].get(site, 0) + 1


def run_classification(config: dict, limit: int = None):
    """Run classification with given configuration."""
    cls_state = app_state['classification']

    try:
        # Reset state
        cls_state.update({
            'running': True,
            'paused': False,
            'processed': 0,
            'errors': 0,
            'results': [],
            'statistics': {'periods': {}, 'decorations': {}, 'confidences': [], 'sites': {}}
        })

        socketio.emit('status', {'status': 'starting', 'message': 'Connecting to database...'})

        # Connect to database
        db_config = config.get('database', {})
        connector = create_connector(db_config.get('type'), db_config.get('connection'))

        if not connector.connect():
            socketio.emit('status', {'status': 'error', 'message': 'Database connection failed'})
            cls_state['running'] = False
            return

        # Execute query
        query = config.get('query')
        if not query:
            socketio.emit('status', {'status': 'error', 'message': 'No query configured'})
            cls_state['running'] = False
            return

        socketio.emit('status', {'status': 'running', 'message': 'Executing query...'})
        items = connector.execute_query(query)
        connector.disconnect()

        if limit:
            items = items[:limit]

        cls_state['total'] = len(items)
        socketio.emit('status', {
            'status': 'running',
            'message': f'Found {len(items)} items to classify',
            'total': len(items)
        })

        # Get image configuration
        image_config = config.get('images', {})
        base_path = image_config.get('base_path', '')
        pattern = image_config.get('pattern', '{media_id}_{filename}.png')

        # Process each item
        for idx, item in enumerate(items):
            while cls_state['paused'] and cls_state['running']:
                socketio.sleep(0.5)

            if not cls_state['running']:
                break

            result_item = dict(item)

            # Find image
            image_path = find_image_file(base_path, pattern, item)

            if image_path:
                result_item['thumbnail'] = create_thumbnail_base64(image_path)
                result_item['image_path'] = str(image_path)

                classification = classify_image(image_path)
                result_item['classification'] = classification

                if classification.get('success'):
                    update_statistics(classification)
                    result_item['status'] = 'success'
                else:
                    cls_state['errors'] += 1
                    result_item['status'] = 'error'
            else:
                cls_state['errors'] += 1
                result_item['status'] = 'error'
                result_item['classification'] = {'success': False, 'error': 'Image not found'}

            cls_state['results'].append(result_item)
            cls_state['processed'] += 1

            socketio.emit('progress', {
                'processed': cls_state['processed'],
                'total': cls_state['total'],
                'errors': cls_state['errors'],
                'current_item': result_item,
                'statistics': cls_state['statistics']
            })

            socketio.sleep(0.1)

        # Complete
        cls_state['running'] = False
        socketio.emit('status', {
            'status': 'completed',
            'message': f'Classification completed',
            'processed': cls_state['processed'],
            'errors': cls_state['errors']
        })

    except Exception as e:
        cls_state['running'] = False
        socketio.emit('status', {'status': 'error', 'message': str(e)})
        traceback.print_exc()


def export_to_excel():
    """Export results to Excel with thumbnails."""
    results = app_state['classification']['results']
    if not results:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Classification Results"

    # Get all unique keys from results
    all_keys = set()
    for r in results:
        all_keys.update(r.keys())

    # Remove internal keys
    exclude_keys = {'thumbnail', 'image_path', 'classification', 'status'}
    data_keys = sorted([k for k in all_keys if k not in exclude_keys])

    headers = ['Thumbnail'] + data_keys + ['ML Period', 'ML Confidence', 'ML Decoration',
                                            'Top Match', 'Similarity', 'Status']

    # Write headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    ws.column_dimensions['A'].width = 12
    temp_files = []

    for row_num, item in enumerate(results, 2):
        ws.row_dimensions[row_num].height = 60

        # Add thumbnail
        thumbnail = item.get('thumbnail')
        if thumbnail and thumbnail.startswith('data:image'):
            try:
                b64_data = thumbnail.split(',')[1]
                img_bytes = base64.b64decode(b64_data)
                temp_path = tempfile.mktemp(suffix='.jpg')
                with open(temp_path, 'wb') as f:
                    f.write(img_bytes)
                temp_files.append(temp_path)

                img = XLImage(temp_path)
                img.width = 60
                img.height = 60
                ws.add_image(img, f'A{row_num}')
            except:
                pass

        classification = item.get('classification', {})
        top_match = classification.get('top_match', {}) or {}

        row_data = ['']  # Thumbnail placeholder
        for key in data_keys:
            row_data.append(item.get(key, ''))

        row_data.extend([
            classification.get('period_suggestion', ''),
            classification.get('period_confidence', ''),
            classification.get('decoration_suggestion', ''),
            top_match.get('id', ''),
            f"{top_match.get('similarity', 0):.1f}%" if top_match.get('similarity') else '',
            item.get('status', '')
        ])

        for col, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col, value=value)

    ws.freeze_panes = 'A2'

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = EXPORT_DIR / f"classification_{timestamp}.xlsx"
    wb.save(output_path)

    for f in temp_files:
        try:
            os.remove(f)
        except:
            pass

    return str(output_path)


# ============== API Routes ==============

@app.route('/api/status')
def api_status():
    cls = app_state['classification']
    return jsonify({
        'running': cls['running'],
        'paused': cls['paused'],
        'total': cls['total'],
        'processed': cls['processed'],
        'errors': cls['errors'],
        'statistics': cls['statistics']
    })


@app.route('/api/export')
def api_export():
    path = export_to_excel()
    if path:
        return jsonify({'success': True, 'path': path})
    return jsonify({'success': False, 'error': 'No results'})


@app.route('/api/download')
def api_download():
    # Find latest export
    exports = list(EXPORT_DIR.glob('*.xlsx'))
    if exports:
        latest = max(exports, key=lambda p: p.stat().st_mtime)
        return send_file(latest, as_attachment=True)
    return jsonify({'error': 'No export available'}), 404


# ============== Plate Generator API ==============

PLATES_DIR = BASE_DIR / "exports" / "plates"
PLATES_DIR.mkdir(parents=True, exist_ok=True)


@app.route('/api/plates/periods')
def api_plates_periods():
    """Get available periods from classification results."""
    results = app_state['classification'].get('results', [])
    periods = {}

    for result in results:
        if result.get('status') == 'success':
            classification = result.get('classification', {})
            period = classification.get('period_suggestion', 'Unknown')
            if period:
                periods[period] = periods.get(period, 0) + 1

    return jsonify({
        'success': True,
        'periods': [{'name': k, 'count': v} for k, v in sorted(periods.items())]
    })


@app.route('/api/plates/items')
def api_plates_items():
    """Get classified items filtered by period and/or US."""
    period = request.args.get('period')
    us = request.args.get('us')

    results = app_state['classification'].get('results', [])
    items = []

    for result in results:
        if result.get('status') != 'success':
            continue

        classification = result.get('classification', {})
        item_period = classification.get('period_suggestion', '')

        if period and item_period != period:
            continue

        item_us = result.get('database_fields', {}).get('us', result.get('us', ''))
        if us and str(item_us) != str(us):
            continue

        items.append({
            'id': result.get('id', result.get('database_fields', {}).get('id', 'N/A')),
            'us': item_us,
            'period': item_period,
            'thumbnail': result.get('thumbnail', ''),
            'image_path': result.get('image_path', ''),
            'form': result.get('database_fields', {}).get('form', ''),
            'decoration': result.get('database_fields', {}).get('decoration', '')
        })

    return jsonify({
        'success': True,
        'items': items,
        'count': len(items)
    })


@app.route('/api/plates/layouts')
def api_plates_layouts():
    """Get available layout definitions."""
    return jsonify({
        'success': True,
        'layouts': get_layouts()
    })


@app.route('/api/plates/preview', methods=['POST'])
def api_plates_preview():
    """Generate preview image for a plate."""
    try:
        data = request.json
        item_ids = data.get('items', [])
        layout_id = data.get('layout_id', '2x2')
        caption_format = data.get('caption_format', 'inv')

        # Convert IDs to full item objects from classification results
        results = app_state['classification'].get('results', [])
        items = []
        for result in results:
            if result.get('status') == 'success':
                item_id = result.get('id', result.get('database_fields', {}).get('id', ''))
                if str(item_id) in [str(i) for i in item_ids]:
                    classification = result.get('classification', {})
                    items.append({
                        'id': item_id,
                        'us': result.get('database_fields', {}).get('us', result.get('us', '')),
                        'period': classification.get('period_suggestion', ''),
                        'thumbnail': result.get('thumbnail', ''),
                        'image_path': result.get('image_path', '')
                    })

        generator = PlateGenerator(str(PLATES_DIR))
        preview = generator.generate_preview(items, layout_id, caption_format)

        return jsonify({
            'success': True,
            'preview': preview
        })

    except Exception as e:
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        })


@app.route('/api/plates/generate', methods=['POST'])
def api_plates_generate():
    """Generate PDF plates."""
    try:
        data = request.json
        item_ids = data.get('items', [])
        layout_id = data.get('layout_id', '2x2')
        period = data.get('period', 'Unknown')
        group_by_us = data.get('group_by_us', False)
        caption_format = data.get('caption_format', 'inv')
        start_plate_number = data.get('start_plate_number', 1)

        # Convert IDs to full item objects from classification results
        results = app_state['classification'].get('results', [])
        items = []
        for result in results:
            if result.get('status') == 'success':
                item_id = result.get('id', result.get('database_fields', {}).get('id', ''))
                if str(item_id) in [str(i) for i in item_ids]:
                    classification = result.get('classification', {})
                    items.append({
                        'id': item_id,
                        'us': result.get('database_fields', {}).get('us', result.get('us', '')),
                        'period': classification.get('period_suggestion', ''),
                        'thumbnail': result.get('thumbnail', ''),
                        'image_path': result.get('image_path', '')
                    })

        generator = PlateGenerator(str(PLATES_DIR))
        filename, assignments = generator.generate_plates(
            items=items,
            layout_id=layout_id,
            period=period,
            group_by_us=group_by_us,
            caption_format=caption_format,
            start_plate_number=start_plate_number
        )

        return jsonify({
            'success': True,
            'filename': filename,
            'plates_count': len(set(a['plate_number'] for a in assignments)),
            'items_count': len(assignments)
        })

    except Exception as e:
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        })


@app.route('/api/plates/report', methods=['POST'])
def api_plates_report():
    """Generate Excel report for plate assignments."""
    try:
        data = request.json
        plates = data.get('plates', [])

        # Convert nested plates data to flat assignments format
        assignments = []
        for plate in plates:
            plate_number = plate.get('plate_number', 0)
            period = plate.get('period', '')
            us = plate.get('us', '')
            items = plate.get('items', [])

            for pos, item in enumerate(items, 1):
                assignments.append({
                    'plate_number': plate_number,
                    'position': pos,
                    'id': item.get('id', ''),
                    'us': item.get('us', us),
                    'period': period,
                    'layout': data.get('layout_id', '2x3')
                })

        filename = generate_excel_report(assignments, str(PLATES_DIR))

        return jsonify({
            'success': True,
            'filename': filename
        })

    except Exception as e:
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        })


@app.route('/api/plates/download/<filename>')
def api_plates_download(filename):
    """Download generated plate file."""
    filepath = PLATES_DIR / filename
    if filepath.exists():
        return send_file(filepath, as_attachment=True)
    return jsonify({'error': 'File not found'}), 404


# ============== Socket Events ==============

@socketio.on('connect')
def handle_connect():
    emit('status', {
        'status': 'connected',
        'running': app_state['classification']['running']
    })


@socketio.on('start')
def handle_start(data):
    if app_state['classification']['running']:
        emit('status', {'status': 'error', 'message': 'Already running'})
        return

    config = data.get('config') or app_state.get('config')
    if not config:
        emit('status', {'status': 'error', 'message': 'No configuration'})
        return

    limit = data.get('limit')
    socketio.start_background_task(run_classification, config, limit)


@socketio.on('pause')
def handle_pause():
    cls = app_state['classification']
    cls['paused'] = not cls['paused']
    emit('status', {'status': 'paused' if cls['paused'] else 'running'})


@socketio.on('stop')
def handle_stop():
    app_state['classification']['running'] = False
    emit('status', {'status': 'stopped'})


if __name__ == '__main__':
    print("=" * 60)
    print("Ceramica Classifier - Universal System")
    print("=" * 60)
    print(f"Server: http://localhost:5002")
    print(f"ML API: {ML_API_URL}")
    print("=" * 60)

    socketio.run(app, host='0.0.0.0', port=5002, debug=True)
